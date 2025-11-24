import os
folder_name = 'フォルダーテスト'
#os.mkdir(folder_name)

import os
folder_name = 'フォルダーテスト'
#フォルダの存在を確認
if os.path.exists(folder_name):
    print(f'フォルダ"{folder_name}"は存在します')
else:
    print('フォルダ"{folder_name}"は存在しません')

#フォルダの削除
#os.rmdir(folder_name)
if os.path.exists(folder_name):
    print(f'フォルダ"{folder_name}"は存在します')
else:
    print('フォルダ"{folder_name}"は存在しません')

import os
folder_name = 'フォルダテスト'
#os.mkdir(folder_name)

#txt_file = open('C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w_UTF-8.txt','x',encoding = 'UTF-8')
#txt_date = 'こんばんは\n 明日は雨になりそうですね\n'
#txt_file.write(txt_date)
#txt_file.close()

#txt_file = open('C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w_UTF-8.txt','a',encoding = 'UTF-8')
#txt_date = 'おはようございます\n 昨日はいい天気でしたね\n'
#txt_file.write(txt_date)
#txt_file.close()

#txt_file = open('C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w_UTF-8.txt','w',encoding = 'UTF-8')
#txt_date = 'こんにちは\n 明日もいい天気になりそうですね\n'
#txt_file.write(txt_date)
#txt_file.close()

import os
file_path = 'C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w_UTF-8.txt'
if os.path.exists(file_path):
    os.remove(file_path)
    print(f'ファイル"{file_path}"を削除しました')
else:
    print(f'ファイル"{file_path}"は存在しません')

#import csv
#word = 'apple,banana'
#words = word.split(',')
#with open('C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w_UTF-8.csv','w',encoding = 'UTF-8',newline="")as f:
    #writer =csv.writer(f)
    #writer.writerow(words)

#import csv
#with open('C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w_UTF-8.csv','a',encoding = 'UTF-8',newline="")as f:
    #writer =csv.writer(f)
    #writer.writerow(['peach','orange'])

#import csv
#with open('C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w_UTF-8.csv','r',encoding = 'UTF-8',newline="")as f:
    #reader =csv.reader(f)
    #for line in reader:
        #print(line)

import os
file_name = 'C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w_UTF-8.csv'
if os.path.exists(file_name):
    os.remove(file_name)
    print(f"ファイル'{file_name}'を削除しました")
else:
    print(f"ファイル'{file_name}'は存在しません")

#import csv
#with open('C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w_UTF-8.csv','r',encoding = 'UTF-8',newline="")as f:
    #reader =csv.reader(f)
    #for line in reader:
        #print(line)

import csv
word = 'apple,banana'
words = word.split(',')
with open('C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w_UTF-8.csv','w',encoding = 'Shift-JIS',newline="")as f:
    writer =csv.writer(f)
    writer.writerow(words)

import csv
with open('C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w_UTF-8.csv','a',encoding = 'Shift-JIS',newline="")as f:
    writer =csv.writer(f)
    writer.writerow(['peach','orange'])

import csv
with open('C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w_UTF-8.csv','r',encoding = 'Shift-JIS',newline="")as f:
    reader =csv.reader(f)
    for line in reader:
        print(line)

import os
file_name = 'C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w_UTF-8.csv'
if os.path.exists(file_name):
    os.remove(file_name)
    print(f"ファイル'{file_name}'を削除しました")
else:
    print(f"ファイル'{file_name}'は存在しません")

from openpyxl import Workbook
wb = Workbook()
sheet = wb.active
sheet['A1'] = "おはよう"
sheet['B1'] = "こんにちは"
wb.save('C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w.xlsx')

from openpyxl import Workbook,load_workbook
file_path = 'C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w.xlsx'
wb = load_workbook(file_path)
sheet = wb.active
sheet['C1'] = "こんばんは"
wb.save(file_path)

from openpyxl import Workbook,load_workbook
wb = load_workbook('C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト\\test1-w.xlsx')
ws = wb['Sheet']
print(ws.cell(1,1).value)

import shutil
shutil.make_archive('test1-w.xlsx',format="zip",root_dir='C:\\Users\\ooooi\\Desktop\\Python\\フォルダテスト')

import shutil
shutil.unpack_archive('test1-w.xlsx.zip')